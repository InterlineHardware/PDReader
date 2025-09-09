import openpyxl
import os
from urllib.parse import urlparse
from src.Products import PDRow
import src.Logline as Logline
from src.Products import *
from src.Matrixify import *
import src.Logline as Logline

logger = Logline.Logger(__name__).logger

class PDProcessor:
    
    def __init__(self, inputFile: str, inputSheet: str):
        self.inputFile = inputFile
        self.inputSheet = inputSheet
    
    def __str__(self) -> str:
        return f"Filename: {self.inputFile}\nSheet name: {self.inputSheet}"        
    
    def readPDProducts(self, startRow: int, endRow: int) -> list:
        
        #  > Opens and Reads The PD File. Dependant on Header Naming
        #  > Interline SKU
        #  > Manufacturer Model No
        #  > Brand

        try:
            wb = openpyxl.load_workbook(self.inputFile)
            sheet = wb[self.inputSheet]
        except Exception as err:
            logger.error(f"ERROR Opening Excel FIle {self.inputFile} : {err}")
        
        column_headers = [cell.value for cell in sheet[1]]
        header_index = {header: index for index, header in enumerate(column_headers)}
        
        PDproducts :PDRow = []

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, max_row=endRow, values_only=True), start=2):

            #--------------Raw Value Extraction-------------------------
            try:
                partNo = str(row[header_index["Interline SKU"]]).strip()
                status = True if (row[header_index["Status"]]) == 0 else False
                isOnline = bool(row[header_index["Is Online"]])
                isVariant = bool(row[header_index["Is Variant"]]) if "Is Variant" in header_index and row[header_index["Is Variant"]] is not None else False
                modelno = str(row[header_index["Manufacturer Model No"]]).strip() if row[header_index["Manufacturer Model No"]] != None else ""
                mfg_part_no = str(row[header_index["Manufacturer Part No"]]).strip() if row[header_index["Manufacturer Part No"]] != None else ""
                title = str(row[header_index["Title"]]).strip() if row[header_index["Title"]] != None else ""
                
                upc = str(row[header_index["UPC"]]).strip() if "UPC" in header_index and row[header_index["UPC"]] is not None else ""
                brand = str(row[header_index["Brand"]]).strip() if "Brand" in header_index and row[header_index["Brand"]] is not None else ""
                tags = str(row[header_index["Tags"]]).strip() if "Tags" in header_index and row[header_index["Tags"]] is not None else ""

                isPreSale = row[header_index["Is Pre-Sale"]] if "Is Pre-Sale" in header_index and row[header_index["Is Pre-Sale"]] is not None else ""
                continueSelling = row[header_index["Continue Selling Out of Stock"]] if "Continue Selling Out of Stock" in header_index and row[header_index["Continue Selling Out of Stock"]] is not None else ""
                specialDelivery = row[header_index["Special Delivery"]] if "Special Delivery" in header_index and row[header_index["Special Delivery"]] is not None else ""

                InventoryPolicy = 'continue' if continueSelling else 'deny'
                if specialDelivery:
                    tags += ',' + 'special_delivery' if tags != '' else 'special_delivery'
                                
                option1_name = str(row[header_index["Option 1 Name"]]).strip() if "Option 1 Name" in header_index and row[header_index["Option 1 Name"]] is not None else ""
                option1_value = str(row[header_index["Option 1 Value"]]).strip() if "Option 1 Value" in header_index and row[header_index["Option 1 Value"]] is not None else ""
                option2_name = str(row[header_index["Option 2 Name"]]).strip() if "Option 2 Name" in header_index and row[header_index["Option 2 Name"]] is not None else ""
                option2_value = str(row[header_index["Option 2 Value"]]).strip() if "Option 2 Value" in header_index and row[header_index["Option 2 Value"]] is not None else ""
                option3_name = str(row[header_index["Option 3 Name"]]).strip() if "Option 3 Name" in header_index and row[header_index["Option 3 Name"]] is not None else ""
                option3_value = str(row[header_index["Option 3 Value"]]).strip() if "Option 3 Value" in header_index and row[header_index["Option 3 Value"]] is not None else ""
                option4_name = str(row[header_index["Option 4 Name"]]).strip() if "Option 4 Name" in header_index and row[header_index["Option 4 Name"]] is not None else ""
                option4_value = str(row[header_index["Option 4 Value"]]).strip() if "Option 4 Value" in header_index and row[header_index["Option 4 Value"]] is not None else ""

                pack_quantity = str(row[header_index["Pack Quantity"]]).strip() if "Pack Quantity" in header_index and row[header_index["Pack Quantity"]] is not None else ""

                category1 = str(row[header_index["Category 1"]]).strip() if "Category 1" in header_index and row[header_index["Category 1"]] is not None else ""
                category2 = str(row[header_index["Category 2"]]).strip() if "Category 2" in header_index and row[header_index["Category 2"]] is not None else ""
                category3 = str(row[header_index["Category 3"]]).strip() if "Category 3" in header_index and row[header_index["Category 3"]] is not None else ""
                category4 = str(row[header_index["Category 4"]]).strip() if "Category 4" in header_index and row[header_index["Category 4"]] is not None else ""
                battery_system = str(row[header_index["Battery System"]]).strip() if "Battery System" in header_index and row[header_index["Battery System"]] is not None else ""

                variantDescription = str(row[header_index["Variant Description"]]) if "Variant Description" in header_index and row[header_index["Variant Description"]] is not None else ""
                description = str(row[header_index["Description"]]).strip() if "Description" in header_index and row[header_index["Description"]] is not None else ""
                html_desc_features = str(row[header_index["HTML Desc/Features"]]).strip() if "HTML Desc/Features" in header_index and row[header_index["HTML Desc/Features"]] is not None else ""
                json_includes = str(row[header_index["JSON Includes"]]).strip() if "JSON Includes" in header_index and row[header_index["JSON Includes"]] is not None else ""
                json_applications = str(row[header_index["JSON Applications"]]).strip() if "JSON Applications" in header_index and row[header_index["JSON Applications"]] is not None else ""
                json_specs = str(row[header_index["JSON Specs"]]).strip() if "JSON Specs" in header_index and row[header_index["JSON Specs"]] is not None else ""
                concatenated_features = str(row[header_index["Concatenated Features"]]).strip() if "Concatenated Features" in header_index and row[header_index["Concatenated Features"]] is not None else ""
                concatenated_includes = str(row[header_index["Concatenated Includes"]]).strip() if "Concatenated Includes" in header_index and row[header_index["Concatenated Includes"]] is not None else ""
                concatenated_specs = str(row[header_index["Concatenated Specs"]]).strip() if "Concatenated Specs" in header_index and row[header_index["Concatenated Specs"]] is not None else ""
                concatenated_applications = str(row[header_index["Concatenated Applications"]]).strip() if "Concatenated Applications" in header_index and row[header_index["Concatenated Applications"]] is not None else ""
                images = str(row[header_index["Concatenated Images"]]).strip() if "Concatenated Images" in header_index and row[header_index["Concatenated Images"]] is not None else ""
                videos = str(row[header_index["Product Video"]]).strip() if "Product Video" in header_index and row[header_index["Product Video"]] is not None else ""

                more_info_url = str(row[header_index["More Info URL"]]).strip() if "More Info URL" in header_index and row[header_index["More Info URL"]] is not None else ""
                tds = str(row[header_index["Technical Data Sheet"]]).strip() if "Technical Data Sheet" in header_index and row[header_index["Technical Data Sheet"]] is not None else ""
                sds = str(row[header_index["Safety Data Sheet"]]).strip() if "Safety Data Sheet" in header_index and row[header_index["Safety Data Sheet"]] is not None else ""
                spec_sheet = str(row[header_index["Spec Sheet"]]).strip() if "Spec Sheet" in header_index and row[header_index["Spec Sheet"]] is not None else ""
                alt_res_name = str(row[header_index["Alternate Resource Name"]]).strip() if "Alternate Resource Name" in header_index and row[header_index["Alternate Resource Name"]] is not None else ""
                alt_res_value = str(row[header_index["Alternate Resource Value"]]).strip() if "Alternate Resource Value" in header_index and row[header_index["Alternate Resource Value"]] is not None else ""

                assembled_weight = str(row[header_index["Assembled Weight"]]).strip() if "Assembled Weight" in header_index and row[header_index["Assembled Weight"]] is not None else ""
                assembled_length = str(row[header_index["Assembled Length"]]).strip() if "Assembled Length" in header_index and row[header_index["Assembled Length"]] is not None else ""
                assembled_width = str(row[header_index["Assembled Width"]]).strip() if "Assembled Width" in header_index and row[header_index["Assembled Width"]] is not None else ""
                assembled_height = str(row[header_index["Assembled Height"]]).strip() if "Assembled Height" in header_index and row[header_index["Assembled Height"]] is not None else ""
                assembled_depth = str(row[header_index["Assembled Depth"]]).strip() if "Assembled Depth" in header_index and row[header_index["Assembled Depth"]] is not None else ""
                assembled_operating_weight = str(row[header_index["Assembled Operating Weight"]]).strip() if "Assembled Operating Weight" in header_index and row[header_index["Assembled Operating Weight"]] is not None else ""
                
                packaged_weight = str(row[header_index["Packaged Weight"]]).strip() if "Packaged Weight" in header_index and row[header_index["Packaged Weight"]] is not None else ""
                packaged_length = str(row[header_index["Packaged Length"]]).strip() if "Packaged Length" in header_index and row[header_index["Packaged Length"]] is not None else ""
                packaged_width = str(row[header_index["Packaged Width"]]).strip() if "Packaged Width" in header_index and row[header_index["Packaged Width"]] is not None else ""
                packaged_height = str(row[header_index["Packaged Height"]]).strip() if "Packaged Height" in header_index and row[header_index["Packaged Height"]] is not None else ""
                packaged_depth = str(row[header_index["Packaged Depth"]]).strip() if "Packaged Depth" in header_index and row[header_index["Packaged Depth"]] is not None else ""
                volume = str(row[header_index["Volume"]]).strip() if "Volume" in header_index and row[header_index["Volume"]] is not None else ""
            
            except KeyError as e:
                logger.error(f"Missing column in PD file: {e}")
                raise KeyError(f"Column {e} not found in the PD file. Please check the column headers.")
        #--------------------PD ROW Object Creation---------------------------

            product = PDRow(
                interline_sku=partNo, manufacturer_model_no=modelno, upc=upc, manufacturer_part_no=mfg_part_no,
                status=status, is_online=isOnline, brand=brand, title=title, is_variant=isVariant,
                option1_name=option1_name, option1_value=option1_value, option2_name=option2_name, option2_value=option2_value,
                option3_name=option3_name, option3_value=option3_value, option4_name=option4_name, option4_value=option4_value,
                pack_quantity=pack_quantity, category1=category1, category2=category2, category3=category3, category4=category4,
                tags=tags, battery_system=battery_system,description=description, html_desc_features=html_desc_features, json_includes=json_includes,
                json_applications=json_applications, json_specs=json_specs, images=images, videos=videos, more_info_url=more_info_url,
                tds=tds, sds=sds, spec_sheet=spec_sheet, alt_res_name=alt_res_name, alt_res_value=alt_res_value,
                assembled_weight=assembled_weight, packaged_weight=packaged_weight,
                concatenated_applications=concatenated_applications, concatenated_features=concatenated_features, concatenated_specs=concatenated_specs,
                concatenated_includes=concatenated_includes, assembled_length=assembled_length, assembled_depth=assembled_depth,assembled_height=assembled_height,
                assembled_width=assembled_width,assembled_operating_weight=assembled_operating_weight ,packaged_depth=packaged_depth, packaged_height=packaged_height,
                packaged_length=packaged_length,packaged_width=packaged_width, volume=volume,preSale=isPreSale,InventoryPolicy=InventoryPolicy,variant_description=variantDescription
            )
            PDproducts.append(product)
        logger.info(f"{len(PDproducts)} Rows Read")
        errors = self.validate_variants(PDproducts)
        if errors:
            logger.info("Errors found in Product Data:\n" + "\n".join([f"{idx+1}. {err}" for idx, err in enumerate(errors)]))
            raise ValueError()
        return PDproducts
    
    def validate_variants(self, pdrows):
        """
        Validates:
        1. If isVariant is True, all products with same title are variants.
        2. For each group of variants, the combination of variant option values is unique.
        3. For each group of variants, html_desc_features are identical.

        Returns:
            List of validation error messages.
        """
        errors = []
        variant_groups = {}

        # Group products by title where isVariant is True
        for row in pdrows:
            if row.isVariant:
                variant_groups.setdefault(row.title, []).append(row)

        for title, variants in variant_groups.items():
            # Validate html_desc_features consistency
            html_desc_set = set(v.html_desc_features for v in variants)
            if len(html_desc_set) > 1:
                errors.append(f"html_desc_features mismatch in variants with title '{title}'.")

            # Validate uniqueness of combination of variant option values
            combination_set = set()
            for v in variants:
                opts = v.variantProperties
                combination = (
                    opts.option1.value or "",
                    opts.option2.value or "",
                    opts.option3.value or "",
                    opts.option4.value or ""
                )
                if combination in combination_set:
                    errors.append(f"Duplicate variant option combination {combination} for title '{title}'.")
                else:
                    combination_set.add(combination)

        return errors


    def generateMatrixifyFile(self, products: list[Product]):
        
        #We want to initialize the output file first
        #Set the column headers using Matrixify 'columns' attribute

        wb = openpyxl.Workbook()
        
        sheet = wb.active
        sheet.title = "Products"
        sheet.append(Matrixify.columns)

        #Start looping
        for product in products:
            row = Matrixify.ProductToRow(product).row
            sheet.append(row)
        
        logger.info(f"{len(products)} Matrixify Rows Generated")
        
        file_name = os.path.basename(self.inputFile)
        file = "Matrixify/" + file_name.replace("Stage 2.xlsx", "") + "Matrixify.xlsx"
        
        wb.save(file)
        logger.info(f"Matrixify File Generated : {file}")

   #-----------------------------------------------------------------------------------------------------------#
   #-----------------------------------------------------------------------------------------------------------#
   #-----------------------------------------------------------------------------------------------------------#
   #-----------------------------------------------------------------------------------------------------------#
   #-----------------------------------------------------------------------------------------------------------#
   # ----------------------------------------------------------------------------------------------------------#
    
    def writeLinks(self, mxRows, maxRow):

    # --> Takes in  a list of Matrixify Files Sheet Row objects
    # --> Reads Links
    # --> Reads Partnumber and File Type from ALT TEXT COlLUMN
    # --> Writes the file link to the appropriate Column IN PD FILE


        wb = openpyxl.load_workbook(self.inputFile)
        sheet = wb[self.inputSheet]
        logger.info(f"{self.inputFile} Opened to Write Links")

        for mxRow in mxRows:
            link = mxRow.link

            models = "-".join(mxRow.alt_text.split("-")[1:-1]).split("&")
            print(mxRow.alt_text)
            file_type = mxRow.alt_text.split("-")[-1]

            # Read column headers
            column_headers = [cell.value for cell in sheet[1]]
            header_index = {header: index for index, header in enumerate(column_headers)}

            # Determine the correct column index for the file type
            if file_type == "Technical Data Sheet":
                file_type_column = header_index.get("Technical Data Sheet") + 1
            elif file_type == "Safety Data Sheet":
                file_type_column = header_index.get("Safety Data Sheet") + 1
            elif file_type == "Spec Sheet":
                file_type_column = header_index.get("Spec Sheet") + 1
            else:
                # For alternate resource columns
                alt_res_name_col = header_index.get("Alternate Resource Name") + 1
                alt_res_value_col = header_index.get("Alternate Resource Value") + 1

            # Update the cells based on the model numbers
            for model in models:
                print(models)
                # Find the row number for the manufacturer model number
                row_index = self.find_row(sheet, header_index.get("Manufacturer Model No") + 1, model, maxRow)
                if row_index:
                    print(f"Writing {file_type} Link for model {model}")
                    if file_type in ["Technical Data Sheet", "Safety Data Sheet", "Spec Sheet"]:
                        sheet.cell(row=row_index, column=file_type_column).value = link
                    else:
                        sheet.cell(row=row_index, column=alt_res_name_col).value = file_type
                        sheet.cell(row=row_index, column=alt_res_value_col).value = link
        
        wb.save(self.inputFile)
        logger.info(f"{self.inputFile} Saved after Writing links")

    def find_row(self, sheet, column_index, value, maxRow):
        for row in sheet.iter_rows(min_row=2, max_row=maxRow, min_col=column_index, max_col=column_index):
            cell = row[0]
            if cell.value == value:
                return cell.row
        return None 

