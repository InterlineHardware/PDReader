import openpyxl
import src.Logline as Logline
import os
import json

#=============================================================================================================================================================#
#=============================================================================================================================================================#

description = 'AA'
feat_start = 'AB'
feat_end = 'AN'
concatFeat = 'AO'
HTMLFeatDesc = 'AP'

inc_start = 'AQ'
inc_end = 'AZ'
concatInc = 'BA'
JSONInc = 'BB'

appStart = 'MM'
appEnd = 'MM'
concatApp = 'MM'
JSONApp = 'MM'

specStart = 'BC'
specEnd = 'CR'
concatSpec = 'CS'
JSONSpecs = 'CT'

imgStart = 'CU'
imgEnd = 'DI'
concatImgs = 'DJ'



#=============================================================================================================================================================#
#=============================================================================================================================================================#

logger = Logline.Logger("Stage 2 Convertor").logger


def col_to_num(col_str):
    """Convert an Excel column label to its numerical index (1-based)."""
    num = 0
    for char in col_str:
        num = num * 26 + ord(char.upper()) - ord('A') + 1
    return num

def get_formatted_description_and_features(ws, row_num, desc_col, start_feature_col, end_feature_col):
    """Get formatted description and features for a given row."""
    # Fetch and format description from the specified column
    description = ws.cell(row=row_num, column=col_to_num(desc_col)).value
    formatted_desc = f"<p>{description.strip()}</p>\n" if description else ""

    # Collect and format feature bullets
    feature_bullets = []
    for col in range(col_to_num(start_feature_col), col_to_num(end_feature_col) + 1):
        cell_value = ws.cell(row=row_num, column=col).value
        if cell_value:
            feature_bullets.append(f"<li>{cell_value.strip()}</li>\n")
    formatted_features = f"<ul>\n{''.join(feature_bullets)}</ul>" if feature_bullets else ""

    # Combine description and features
    return formatted_desc + formatted_features


def get_formatted_attributes(ws, row_num, start_attr_col, end_attr_col):
    """Returns a JSON Dictionary where specname is the key, and specvalue is the value."""
    specs = {}
    for col in range(col_to_num(start_attr_col), col_to_num(end_attr_col) + 1, 1):
        # Fetch attribute title and value pairs
        title = ws.cell(row=1, column=col).value or ""
        value = ws.cell(row=row_num, column=col).value or ""
        if title in specs:
            logger.warning(f"Duplicate spec found in Row No : {row_num}")

        specs[title] = value

    return json.dumps(specs)

def get_concatenated_attributes(ws, row_num, start_attr_col, end_attr_col):
    """Concatenated Specifications with spec name as column header and spec value in column."""
    attributes = []
    for col in range(col_to_num(start_attr_col), col_to_num(end_attr_col) + 1, 1):
        # Fetch attribute title and value pairs
        title = ws.cell(row=1, column=col).value or ""
        value = ws.cell(row=row_num, column=col).value or ""
        if value != "":
            attributes.append(str(title).strip() + ' : ' + str(value).strip())
    
    return '\n'.join(attributes) if attributes else ""

def get_formatted_attributes2(ws, row_num, start_attr_col, end_attr_col):
    """Returns a JSON Dictionary where specname is the key, and specvalue is the value"""
    specs = {}
    for col in range(col_to_num(start_attr_col), col_to_num(end_attr_col) + 1, 2):
        # Fetch attribute title and value pairs
        title = ws.cell(row=row_num, column=col).value or ""
        value = ws.cell(row=row_num, column=col + 1).value or ""
        if value == '':
            continue
        if title in specs:
            logger.warning(f"Duplicate spec found in Row No : {row_num}")

        specs[title] = value
    
    if specs:
        # for key, value in specs.items():
        #     print(f"Key: {key}, Value: {value}, Type: {type(value)}")
        return json.dumps(specs)
    

def get_concatenated_attributes2(ws, row_num, start_attr_col, end_attr_col):
    """Concatenated Specifications with spec name and spec value columns beside each other."""
    attributes = []
    for col in range(col_to_num(start_attr_col), col_to_num(end_attr_col) + 1, 2):
        # Fetch attribute title and value pairs
        title = ws.cell(row=row_num, column=col).value or ""
        value = ws.cell(row=row_num, column=col + 1).value or ""
        if value != "":
            attributes.append(str(title).strip() + ' : ' + str(value).strip())
    
    return '\n'.join(attributes) if attributes else ""


def get_concatenated(ws, row_num, start_img_col, end_img_col, concatenator):
    """Concatenate cells together from different columns."""
    images = [
    str(ws.cell(row=row_num, column=col).value).strip()
    for col in range(col_to_num(start_img_col), col_to_num(end_img_col) + 1)
    if ws.cell(row=row_num, column=col).value is not None
    ]
    return concatenator.join(images)



def get_formatted_html_list(ws, row_num, start_inc_col, end_inc_col, output_col):
    """Generate and write an HTML list from specified columns to the output column."""
    includes = []

    for col in range(col_to_num(start_inc_col), col_to_num(end_inc_col) + 1):
        cell_value = ws.cell(row=row_num, column=col).value
        if cell_value:  # Only consider non-empty cells
            includes.append(f"<li>{cell_value.strip()}</li>\n")

    if includes:
        ws.cell(row=row_num, column=col_to_num(output_col)).value = f"<ul>\n{''.join(includes)}</ul>"

def delete_multiple_columns(ws, ranges):
    """Delete multiple column ranges in a given worksheet."""
    # Convert all ranges to column numbers and sort in descending order
    col_ranges = []
    for start_col, end_col in ranges:
        start_col_num = col_to_num(start_col)
        end_col_num = col_to_num(end_col)
        col_ranges.extend(range(start_col_num, end_col_num + 1))
    
    col_ranges = sorted(set(col_ranges), reverse=True)
    
    for col_num in col_ranges:
        ws.delete_cols(col_num)

def getRichTextBulletListJSON(ws, row_num, start_inc_col, end_inc_col, output_col):
    list = []

    for col in range(col_to_num(start_inc_col), col_to_num(end_inc_col) + 1):
        cell_value = ws.cell(row=row_num, column=col).value
        if cell_value:  # Only consider non-empty cells
            list.append(str(cell_value).strip())
    rich_text_bulleted_list = {
    "type": "root",
    "children": [
        {
            "listType": "unordered",
            "type": "list",
            "children": [
                {
                    "type": "list-item",
                    "children": [{"type": "text", "value": item}]
                } for item in list
            ]
        }
    ]}

    if list:
        ws.cell(row=row_num, column=col_to_num(output_col)).value = json.dumps(rich_text_bulleted_list) 

    

def main(filename, sheetname, specFormat , data = None):
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheetname]
    logger.info(f"{filename} Opened For Stage 2 Conversion")

    """
    This logic creates new(Updates when variable already exists) Variables from the data form.
    See PDUI For Form Format and Variable Names.
    Keep Variable Name in Form Same as the Variables above in this File To Avoid Compile Errors
    """
    if data is not None:    
        for key, value in data.items():
            globals()[key] = value

    # Iterate over each row and apply transformations


    for row in range(2, ws.max_row + 1):
        
        # Format and write description and features to column
        formatted_desc_features = get_formatted_description_and_features(ws, row, description, feat_start, feat_end) # First letter is description column, second & third are the range of feature bullets
        if formatted_desc_features:
            ws.cell(row=row, column=col_to_num(HTMLFeatDesc)).value = formatted_desc_features # Output Column

        try:
            if specFormat == 2:
                # Format and write specifications to a specific column
                formatted_attrs = get_formatted_attributes2(ws, row, specStart, specEnd) # Range of columns to add to attributes HTML table
                if formatted_attrs:
                    ws.cell(row=row, column=col_to_num(JSONSpecs)).value = formatted_attrs # Output Column

                # Concatenate and write specifications to a specific column
                formatted_attrs = get_concatenated_attributes2(ws, row, specStart, specEnd) # Range of columns to add to attributes concatenation
                if formatted_attrs:
                    ws.cell(row=row, column=col_to_num(concatSpec)).value = formatted_attrs # Output Column
            
            elif specFormat == 1:
                
                # Format and write specifications to a specific column
                formatted_attrs = get_formatted_attributes(ws, row, specStart, specEnd) # Range of columns to add to attributes HTML table
                if formatted_attrs:
                    ws.cell(row=row, column=col_to_num(JSONSpecs)).value = formatted_attrs # Output Column

                # Concatenate and write specifications to a specific column
                formatted_attrs = get_concatenated_attributes(ws, row, specStart, specEnd) # Range of columns to add to attributes concatenation
                if formatted_attrs:
                    ws.cell(row=row, column=col_to_num(JSONSpecs)).value = formatted_attrs # Output Column
        except TypeError as e:
            print(e)
            print(f"Type is not JSON Serializable at row {row}")
            raise(e)
        # Concatenate and write images to a specific column (e.g., 'I')
        concatenated_images = get_concatenated(ws, row, imgStart, imgEnd, ',\n')  # Range of columns with images to concatenate with a ',\n'
        if concatenated_images:
            ws.cell(row=row, column=col_to_num(concatImgs)).value = concatenated_images # Output Column


        #Concatenated Features 
        concatenated_features = get_concatenated(ws, row, feat_start, feat_end, '\n')  # Range of columns with features to concatenate with a '\n'
        if concatenated_features:
            ws.cell(row=row, column=col_to_num(concatFeat)).value = concatenated_features # Output Column
        

        # #Format and write includes html list to specific columns
        # get_formatted_html_list(ws, row, inc_start, inc_end, HTMLInc) # First letters are the range of "includes" points, Last Letter is Output Column
        getRichTextBulletListJSON(ws,row,inc_start,inc_end,JSONInc)

        #Concatenated Includes 
        concatenated_includes = get_concatenated(ws, row, inc_start, inc_end, '\n')  # Range of columns with features to concatenate with a '\n'
        if concatenated_includes:
            ws.cell(row=row, column=col_to_num(concatInc)).value = concatenated_includes # Output Column



        #Concatenated Applications using images function
        concatenated_applications = get_concatenated(ws, row, appStart, appEnd, '\n')  # Range of columns with applications to concatenate with a ','
        if concatenated_applications:
            ws.cell(row=row, column=col_to_num(concatApp)).value = concatenated_applications # Output Column

        #Format and write applications html list to specific columns
        # get_formatted_html_list(ws, row, appStart, appEnd, htmlApp) # First letters are the range of "applications" points, Last Letter is Output Column
        getRichTextBulletListJSON(ws,row,appStart,appEnd,JSONApp)

        #HTML Dimensions
        # html_dimensions = get_formatted_attributes(ws, row, dimStart, dimEnd)
        # if html_dimensions:
        #     ws.cell(row=row, column=col_to_num(HTMLDim)).value = html_dimensions
    
    
    columns_to_delete = [(feat_start, feat_end), (inc_start, inc_end), (appStart, appEnd), (specStart, specEnd), (imgStart, imgEnd) ]
    
    # Delete columns
    delete_multiple_columns(ws, columns_to_delete)
    # Save the modified workbook
    filename = os.path.basename(filename)
    os.makedirs("Stage 2", exist_ok=True)
    wb.save("Stage 2/"+ filename.replace(".xlsx", " - Stage 2.xlsx")  )

    logger.info(F"{filename} Stage 2 Conversion Success")

#=============================================================================================================================================================#
#=============================================================================================================================================================#

if __name__ == "__main__":

    excel_file = "P:/Sanjid & Eric/To Put Online/Milwaukee PD 02-12-2025.xlsx" # Enter excel file name here
    sheet = 'Product Data'
    
    specFormat = 2

    # >>  Spec Format for specifying the fomrat for SpecName and SpecValue Arrangement in PD SHEET.
    # >> 1 -> COLOR || SIZE || WEIGHT
    # >> 2 -> SpecName1 || SpecValue1

    main(excel_file, sheet, specFormat)


