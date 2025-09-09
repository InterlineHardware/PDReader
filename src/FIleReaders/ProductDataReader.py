import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import src.processer as processer
import src.Logline as Logline
import openpyxl
from src.Models.Product import PDRow
from src.QueryModule import QueryModule
import json
from datetime import date, datetime

logger = Logline.Logger(__name__).logger


header_mapping = {
    "product_name": "Extracted Title",
    "description": "Description",
    "features": "Features Bullet 1",
    "includes": "Includes Bullet 1",
    "specs": "Specification Name 1",
    "images": "Main Product Image",
    "video": "Product Video",
    "sds": "Safety Data Sheet",
    "upc": "UPC",
    "assembled_length": "Assembled Length",
    "assembled_width": "Assembled Width",
    "assembled_height": "Assembled Height",
    "assembled_depth": "Assembled Depth",
    "assembled_weight": "Assembled Weight",
    "assembled_operating_weight": "Assembled Operating Weight",
    "packaged_length": "Packaged Length",
    "packaged_width": "Packaged Width",
    "packaged_height": "Packaged Height",
    "packaged_depth": "Packaged Depth",
    "packaged_weight": "Packaged Weight",
    "volume": "Volume",
}


def process(excelfile, SheetName , min_row, max_row):
    Logline.LogData.startTimer()
    workbook = openpyxl.load_workbook(excelfile)
    sheet = workbook[SheetName]
    conn = processer.connect()

    if conn is None or not conn.is_connected():
        logger.error("No Connection with Databse")
        return None

    products = read(sheet, conn, min_row, max_row)
    write(sheet, products, min_row, max_row)
    workbook.save(excelfile)
    logger.info("Workbook Saved")


def read(sheet: openpyxl.worksheet.worksheet.Worksheet, conn, min_row, max_row):
    queryModule = QueryModule(conn)
    headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}
    model_no_col_index = headers["Manufacturer Part No"]

    products: PDRow = []

    for rowNumber, row in enumerate(
        sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True), start=2
    ):
        model_no = row[model_no_col_index - 1]  # Adjust for 0-based index
        if model_no:
            print(f"Row {rowNumber}: Model No = {model_no}")
            product = processer.getProduct(model_no, queryModule)
            if product is not None:
                products.append(product)
    return products


def convert(product: PDRow):
    if product.features is not None:
        features = json.loads(product.features)
        features = "^".join(features)
        product.features = features

    if product.includes is not None:
        includes = json.loads(product.includes)
        includes = "^".join(includes)
        product.includes = includes

    if product.images is not None:
        images = json.loads(product.images)
        images = "^".join(images)
        product.images = images

    if product.specs is not None:
        specs = json.loads(product.specs)
        specs = "^".join(f"{key}^{value}" for key, value in specs.items())
        product.specs = specs

    if product.show_online_date is not None and product.show_online_date > date.today():
        product.product_name = product.product_name + " ***PRE-SALE***"

    return product


def write(sheet: openpyxl.worksheet.worksheet.Worksheet, products: list[PDRow], min_row, max_row):
    headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}
    model_no_col_index = headers.get("Manufacturer Part No")

    if model_no_col_index is None:
        logger.error("Model Number column not found")
        return

    for product in products:
        model_no = product.modelno
        product = convert(product)
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=False):
            if row[model_no_col_index - 1].value == model_no:
                for attribute, header in header_mapping.items():
                    col_index = headers.get(header)
                    if col_index:
                        value = getattr(product, attribute, None)
                        if value is not None:
                            sheet.cell(row=row[0].row, column=col_index).value = value
                break
