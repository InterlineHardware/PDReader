import dateutil.parser
import openpyxl
import dateutil
from datetime import datetime
import json
import src.Logline as Logline
from src.QueryModule import QueryModule
from src.Models.InventoryRecord import InventoryRecord
from src.Models.ProductDetailsRecord import ProductDetailsRecord
from src.Models.UOMRecord import UOMRecord
import src.processer as processer

HEADER_MAP = {
    "MFG_Part_Number": [
        "MFG Part # (OEM)",
        "MFG Part #",
        "Model Number",
        "MFG Model Number",
        "Part Number",
        "SKU",
        "Product SKU",
        "Product Number"
    ],
    "GTIN": ["GTIN", "Global Trade Item Number", "UPC", "EAN"],
    "Short_Description": ["Short Description", "Description Short", "Short Desc"],
    "Product_Name": ["Product Name", "Name", "Item Name"],
    "Search_Keywords": ["Search Keywords", "Keywords"],
}

logger = Logline.Logger(__name__).logger

# >> Excel Processing For Reading A Milwaukee Product Information Fiile


def openFile(excelfile):
    workbook = openpyxl.open(excelfile)
    logger.info("FILE OPENED : " + excelfile)
    return workbook


def get_headers(sheet):
    header_row = sheet[1]  # Assuming headers are in the first row
    headers = {}
    for idx, cell in enumerate(header_row):
        if cell.value:  # Only add non-empty headers
            headers[cell.value] = idx

    return headers

def find_header_index(headers: dict, canonical_name: str):
    """Return the index of the first matching synonym for canonical header key."""
    if canonical_name not in HEADER_MAP:
        raise KeyError(f"No header mapping exists for: {canonical_name}")

    synonyms = HEADER_MAP[canonical_name]

    for s in synonyms:
        if s in headers:
            return headers[s]

    return None

# >> Reads Columns for Creating a Inventory Record


def getInventory(row, header_indices):
    try:
        idx = find_header_index(header_indices, "MFG_Part_Number")
        MFG_Part_Number = row[idx] if idx is not None else None
        GTINCode = row[header_indices["GTIN"]]
        status_index = header_indices.get("Status", None)
        Status = (
            row[status_index]
            if status_index is not None and row[status_index]
            else None
        )
        Short_Description = row[header_indices["Short Description"]]
        Brand = "Milwaukee"
        CountryCode = row[header_indices["Country Code - 2 Character code"]]
        subBrand = row[header_indices["Sub-Brand"]]
        warranty = row[header_indices["Manufacturer Warranty"]]
        recalled = row[header_indices["Has Item Been Recalled"]] == "Y"
        previoussku = row[header_indices["Previous MFG Model #"]]
        previousUPC = row[header_indices["Previous UPC"]]
        minOrderQty = row[header_indices["Minimum Order Quantity"]]
        mulOrderQty = row[header_indices["Multiple Order Quantity"]]
        orderUOM = row[header_indices["Order UOM"]]

    except Exception as err:
        logger.error(f"HEADER ERROR in Inventory query :  {err}")
    try:
        show_online_raw = row[header_indices["Show Online Date"]]
        available_to_ship_raw = row[header_indices["Available to Ship Date"]]

        if isinstance(show_online_raw, datetime):
            Show_Online_Date = show_online_raw.strftime("%Y-%m-%d")
        else:
            Show_Online_Date = dateutil.parser.parse(show_online_raw).strftime("%Y-%m-%d")

        if isinstance(available_to_ship_raw, datetime):
            Available_to_Ship_Date = available_to_ship_raw.strftime("%Y-%m-%d")
        else:
            Available_to_Ship_Date = dateutil.parser.parse(available_to_ship_raw).strftime("%Y-%m-%d")
    
    except Exception as err:
        Show_Online_Date = row[header_indices["Show Online Date"]]
        Available_to_Ship_Date = row[header_indices["Available to Ship Date"]]
        logger.error(
            f"Error parsing Dates{Show_Online_Date, Available_to_Ship_Date} for row with with MFG_Part_Num: {MFG_Part_Number}: {err}"
        )

    inventoryRecord = InventoryRecord(
        MFG_Part_Number=MFG_Part_Number,
        GTIN=GTINCode,
        Status=Status,
        Short_Description=Short_Description,
        Show_Online_Date=Show_Online_Date,
        Available_to_Ship_Date=Available_to_Ship_Date,
        Brand=Brand,
        CountryCode=CountryCode,
        SubBrand=subBrand,
        Warranty=warranty,
        Recalled=recalled,
        PreviousModelNo=previoussku,
        PreviousUPC=previousUPC,
        MinOrder=minOrderQty,
        MulOrder=mulOrderQty,
        OrderUOM=orderUOM,
    )

    return inventoryRecord


# ========================================xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx=========================================================#


# >> Creates A Product Detals Record
def getproductInfo(row, header_indices):
    try:
        idx = find_header_index(header_indices, "MFG_Part_Number")
        modelno = row[idx] if idx is not None else None
        search_keywords = row[header_indices["Search Keywords"]]
        product_name = row[header_indices["Product Name"]]
        description = row[header_indices["Marketing Copy"]]
        features_list = [
            row[header_indices[f"Feature - Benefit Bullet {i}"]]
            for i in range(1, 22)
            if header_indices.get(f"Feature - Benefit Bullet {i}")
            and row[header_indices[f"Feature - Benefit Bullet {i}"]]
        ]
        features = json.dumps(features_list)

        # Create the includes JSON object, filtering out any null values or empty strings
        includes_list = [
            row[header_indices["Package Contents"]]
            if header_indices.get("Package Contents")
            and row[header_indices["Package Contents"]]
            else ""
        ]
        includes = json.dumps(includes_list)

        # Extract assembled dimensions and weight
        assembled_depth = (
            f'{row[header_indices["Assembled Depth (in)"]]}"'
            if header_indices.get("Assembled Depth (in)")
            else None
        )
        assembled_width = (
            f'{row[header_indices["Assembled Width (in)"]]}"'
            if header_indices.get("Assembled Width (in)")
            else None
        )
        assembled_height = (
            f'{row[header_indices["Assembled Height (in)"]]}"'
            if header_indices.get("Assembled Height (in)")
            else None
        )
        assembled_weight = (
            f'{row[header_indices["Assembled Weight (lbs)"]]} lbs'
            if header_indices.get("Assembled Weight (lbs)")
            else None
        )

        packaged_depth = (
            f'{row[header_indices["Package Depth (In.)"]]}"'
            if header_indices.get("Package Depth (In.)")
            else None
        )
        packaged_width = (
            f'{row[header_indices["Package Width (In.)"]]}"'
            if header_indices.get("Package Width (In.)")
            else None
        )
        packaged_height = (
            f'{row[header_indices["Package Height (In.)"]]}"'
            if header_indices.get("Package Height (In.)")
            else None
        )
        packaged_weight = (
            f'{row[header_indices["Package Weight (Lb.)"]]} lbs'
            if header_indices.get("Package Weight (Lb.)")
            else None
        )

    except Exception as err:
        logger.error(f"HEADER ERROR in productInfo query: {err}")
        return None

    # Create ProductDetailsRecord with new columns
    productDetailsRecord = ProductDetailsRecord(
        modelno=modelno,
        search_keywords=search_keywords,
        product_name=product_name,
        description=description,
        features=features,
        includes=includes,
        assembled_depth=assembled_depth,
        assembled_width=assembled_width,
        assembled_height=assembled_height,
        assembled_weight=assembled_weight,
        packaged_depth=packaged_depth,
        packaged_width=packaged_width,
        packaged_height=packaged_height,
        packaged_weight=packaged_weight,
    )

    return productDetailsRecord

    # ========================================xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx=========================================================#


def getDigitalAssets(row, digitalassetsheader):
    try:
        idx = find_header_index(digitalassetsheader, "MFG_Part_Number")
        modelno = row[idx] if idx is not None else None
        video = row[digitalassetsheader["Product Review Video"]]
        safetyDataSheet = row[digitalassetsheader["Safety Data Sheet (SDS) - PDF"]]

        # Find index for the first hero image
        imageStart = digitalassetsheader.get("Main Product Image")
        if imageStart is None:
            raise KeyError("Main Product Image header is missing")

        # Find the index for the last Detailed Product View image
        detailed_views_indices = [
            idx
            for header, idx in digitalassetsheader.items()
            if header.startswith("Detailed Product View")
        ]
        if not detailed_views_indices:
            raise KeyError("No Detailed Product View headers found")
        imagesEnd = max(detailed_views_indices)
    except Exception as err:
        logger.error(f"HEADER ERROR in Digital Assets query : {err}")
    # Extract all images between the first hero image and the last detailed product view
    imagesList = row[imageStart : imagesEnd + 1]
    imagesList = [img for img in imagesList if img]

    # Construct the JSON object
    images = json.dumps(imagesList)

    productDetailsRecord = ProductDetailsRecord(
        modelno=modelno, video=video, sds=safetyDataSheet, images=images
    )
    return productDetailsRecord

    # ========================================xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx=========================================================#


def getUOMs(row, header_indices):
    try:
        idx = find_header_index(header_indices, "MFG_Part_Number")
        mfg_part_no = row[idx] if idx is not None else None
        desc = row[header_indices["Short Description"]]
        uoms = []
        Eachupc = row[header_indices["UPC"]]
        package_height = row[header_indices["Package Height (In.)"]]
        package_width = row[header_indices["Package Width (In.)"]]
        package_depth = row[header_indices["Package Depth (In.)"]]
        package_weight = row[header_indices["Package Weight (Lb.)"]]
        try:
            package_quantity = row[header_indices["Net Package Quantity/Net Content"]]
        except Exception:
            package_quantity = 1
        uom = "EA" if package_quantity == "1" or package_quantity == None else f"PK{package_quantity}"

        EachUOM = UOMRecord(
            mfg_part_no=mfg_part_no,
            desc=desc,
            uom=uom,
            upc=Eachupc,
            quantity=package_quantity,
            weight=package_weight,
            width=package_width,
            height=package_height,
            depth=package_depth,
        )
        uoms.append(EachUOM)
        # Inner pack details
        inner_pack_upc = row[header_indices["Inner Pack GTIN"]]
        inner_pack_quantity = row[header_indices["Inner Pack Quantity"]]
        inner_pack_height = row[header_indices["Inner Pack Height (In.)"]]
        inner_pack_width = row[header_indices["Inner Pack Width (In.)"]]
        inner_pack_depth = row[header_indices["Inner Pack Depth (In.)"]]
        inner_pack_weight = row[header_indices["Inner Pack Weight (Lb.)"]]

        if inner_pack_quantity:
            IpUom = UOMRecord(
                mfg_part_no=mfg_part_no,
                desc=desc,
                quantity=inner_pack_quantity,
                upc=inner_pack_upc,
                uom=f"IP{inner_pack_quantity}",
                height=inner_pack_height,
                depth=inner_pack_depth,
                width=inner_pack_width,
                weight=inner_pack_weight,
            )
            uoms.append(IpUom)
        # Case details
        case_upc = row[header_indices["Case GTIN"]]
        case_quantity = row[header_indices["Case Quantity"]]
        case_height = row[header_indices["Case Height (In.)"]]
        case_width = row[header_indices["Case Width (In.)"]]
        case_depth = row[header_indices["Case Depth (In.)"]]
        case_weight = row[header_indices["Case Weight (Lb.)"]]

        if case_quantity:
            caseUOM = UOMRecord(
                mfg_part_no=mfg_part_no,
                desc=desc,
                quantity=case_quantity,
                uom=f"CASE{case_quantity}",
                upc=case_upc,
                weight=case_weight,
                height=case_height,
                width=case_width,
                depth=case_depth,
            )
            uoms.append(caseUOM)

        return uoms
    except Exception as err:
        logger.error(f"Error Reading Uoms : {err}")

        # ========================================xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx=========================================================#


def getSpecs(row, header_indices):
    try:
        index = header_indices.get("MFG Part # (OEM)")
        if index is not None:
            MFG_Part_Number = row[index]
        else:
            index = header_indices.get("SKU")
            if index is not None:
                MFG_Part_Number = row[index]
            else:
                MFG_Part_Number = None
    except (KeyError, IndexError) as e:
        print(f"Error retrieving part number in Spec Sheet: {e}")
        MFG_Part_Number = None
    
    specDict = {}
    for header_name, index in header_indices.items():
        # Add the header name and value to the dictionary
        if row[index] is None:
            continue
        specDict[header_name] = row[index]

    specDict.pop("GTIN", None)
    specs = json.dumps(specDict)

    productDetails = ProductDetailsRecord(modelno=MFG_Part_Number, specs=specs)
    return productDetails

    # ========================================xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx=========================================================#


def productInformationSheet(workbook, conn):
    logger.info("READING PRODUCT INFORMATION SHEET")
    try:
        productInformationSheet = workbook["Product Information"]
    except Exception:
        productInformationSheet = workbook["Product Information "]
    # Create a dictionary to map header names to column indices
    header_indices = get_headers(productInformationSheet)
    queryModule = QueryModule(conn)

    for row in productInformationSheet.iter_rows(
        min_row=2, max_row=productInformationSheet.max_row, values_only=True
    ):
        idx = find_header_index(header_indices, "MFG_Part_Number")
        MFG_Part_Number = row[idx] if idx is not None else None
        if MFG_Part_Number is None:
            continue
        try:
            # Read Inventory values and run inventory query
            inventoryRecord = getInventory(row, header_indices)
            processer.inventoryQuery(inventoryRecord, queryModule)
            # read product detail values and call productdetails query
            productDetailsRecord = getproductInfo(row, header_indices)
            processer.productInfoQuery(productDetailsRecord, queryModule)
            # read uoms
            uomsRecords = getUOMs(row, header_indices)
            for uom in uomsRecords:
                processer.UOMquery(uom, queryModule)
        except Exception as err:
            logger.error(f"Funciton Error @ProductInfoSheet: {err}")
            Logline.Logger.errortrace(logger, err)
        print(
            "============================000000000000=============================  \n\n\n"
        )

    # ========================================xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx=========================================================


def digitalAssetsSheet(workbook, conn):
    logger.info("READING DIGITAL ASSETS SHEET")
    try:
        digitalAssetsSheet = workbook["Digital Assets"]
    except Exception:
        digitalAssetsSheet = workbook['Digital Assets ']
    digitalAssetsHeaders_indices = get_headers(digitalAssetsSheet)
    queryModule = QueryModule(conn)
    for row in digitalAssetsSheet.iter_rows(
        min_row=2, max_row=digitalAssetsSheet.max_row, values_only=True
    ):
        idx = find_header_index(digitalAssetsHeaders_indices, "MFG_Part_Number")
        MFG_Part_Number = row[idx] if idx is not None else None
        if MFG_Part_Number is None:
            continue
        try:
            # Reads digital Assets and updates/inserts productdetailsrecord
            productDetailsRecord = getDigitalAssets(row, digitalAssetsHeaders_indices)
            processer.productInfoQuery(productDetailsRecord, queryModule)
        except Exception as err:
            logger.error(f"Funciton Error @DigitalAssetsSheet: {err}")
        print(
            "============================000000000000=============================  \n\n\n"
        )

        # ========================================xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx=========================================================#


def specSheets(workbook : openpyxl.Workbook, conn):
    logger.info("READING SPEC SHEETS")
    queryModule = QueryModule(conn)

    excluded_sheets = {
        "Product Information",
        "FR Product Information",
        "Digital Assets",
        "Digital Assets FR",
        'PRODCUT INFORMATION 1',
        'Product Information Tab'

    }
    # print(workbook.sheetnames)
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip() in excluded_sheets:
            continue
        specSheet = workbook[sheet_name]
        # Create a dictionary to map header names to column indices
        specheader_indices = get_headers(specSheet)
        
        # Condition For New Spec Format
        if "SKU" in specheader_indices and "Field Name" in specheader_indices:
            sku_col = specheader_indices["SKU"]
            field_col = specheader_indices["Field Name"]
            value_col = specheader_indices.get("Value", field_col + 1)  # assume value column is next to Field Name if not mapped

            # Dictionary to collect specs per SKU
            products = {}

            for row in specSheet.iter_rows(min_row=2, max_row=specSheet.max_row, values_only=True):
                sku = row[sku_col]
                specName = row[field_col]
                specValue = row[value_col]

                if sku is None or specName is None or specValue in (None, ""):
                    continue

                if sku not in products:
                    products[sku] = {}

                products[sku][specName] = specValue

            # Now create ProductDetailsRecord for each SKU
            for sku, specs_dict in products.items():
                specs_json = json.dumps(specs_dict)
                productDetailsRecord = ProductDetailsRecord(modelno=sku, specs=specs_json)

                try:
                    # print(sku, specs_dict)
                    processer.productInfoQuery(productDetailsRecord, queryModule)
                except Exception as err:
                    logger.error(f"Function Error @SpecSheet:{sheet_name}, SKU:{sku} : {err}")
                    print("============================000000000000=============================")
        
        
        else :
            # Old Format and Master Sheets
            for row in specSheet.iter_rows(
                min_row=2, max_row=specSheet.max_row, values_only=True
            ):
                try:
                    index = specheader_indices.get("MFG Part # (OEM)")
                    if index is not None:
                        MFG_Part_Number = row[index]
                    else:
                        index = specheader_indices.get("SKU")
                        if index is not None:
                            MFG_Part_Number = row[index]
                        else:
                            print(f"Neither 'MFG Part # (OEM)' nor 'SKU' found in sheet {sheet_name}")
                            MFG_Part_Number = None
                except (KeyError, IndexError) as e:
                    print(f"Error retrieving part number in sheet {sheet_name}: {e}")
                    raise Exception
                
                if MFG_Part_Number is None:
                    continue
                try:
                    productDetailsRecord = getSpecs(row, specheader_indices)
                    processer.productInfoQuery(productDetailsRecord, queryModule)
                except Exception as err:
                    logger.error(f"Funciton Error @SpecSheet:{sheet_name} : {err}")
                print(
                    "============================000000000000=============================  \n\n\n"
                )

        # ========================================xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx=========================================================#


def readFile(excelfile):
    Logline.LogData.startTimer()
    workbook = openFile(excelfile)
    conn = processer.connect()

    if conn is None or not conn.is_connected():
        logger.error("No Connection with Databse")
        return None

    productInformationSheet(workbook, conn)
    digitalAssetsSheet(workbook, conn)
    specSheets(workbook, conn)
    conn.close()
